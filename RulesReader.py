# -*- coding: utf-8 -*-
"""
Created on Mon Mar 15 16:32:27 2021

@author: asus
"""

class Rules_administrator(object):
    """This class administrates rules in excel~
    
    Attributes:
        RulesPath: The physical loaction of the rule base
        RulesNum: The total number of rules
        Rules: A list that restore rules temporarily
    """
    def __init__(self, RulesPath):
        self._RulesPath = RulesPath
        self._RulesNum = 0
        self._Rules = list()
        
    def _read_rules(self):
        """Basic function of reading rules
        
        Returns:
            self._Rules(list): A list that restore rules temporarily
        """
        import xlrd
        workbook = xlrd.open_workbook(self._RulesPath)
        content = workbook.sheet_by_index(0)
        self._RulesNum = content.nrows
        for i in range(0, self._RulesNum):
            premise = []
            j = 1
            while j < content.ncols and content.cell_value(i, j) != '':
                premise.append(content.cell_value(i, j))
                j += 1
            conclusion = content.cell_value(i, 0)
            self._Rules.append({conclusion : set(premise)})
        print('Read successfully~')
        return self._Rules
    
    def _show_premises(self):
        """Return a set including all premises
        
        Returns:
            premise(set): all premises
        """
        premise = set()
        for i in range(self._RulesNum):
            premise = premise | list(self._Rules[i].values())[0]
        return premise
    
    def _save_rules(self, NewRulesPath):
        """Store rules in Excel on a path
        
        Args:
            NewRulesPath(str): The path of the excel to save
            
        Returns:
            None
        """
        import openpyxl
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        i = 1
        for rule in self._Rules:
            conclusion = list(rule.keys())[0]
            sheet.cell(row = i, column = 1, value = conclusion)        
            j = 2
            for premise in list(rule.values())[0]:
                sheet.cell(row = i, column = j, value = premise)
                j += 1
            i += 1
        workbook.save(NewRulesPath)
        print('Save successfuly~')
    
    def _if_exist_rule(self, NewRule):
        """Determining whether a rule already existss in the rule base
        
        Args:
            NewRule(dict): The rule to be determined
            
        Returns:
            True(bool): If the rule exists in the rule base
            False(bool): If the rule does not exist in the rule base
        """
        for rule in self._Rules:
            if list(rule.keys())[0] == list(NewRule.keys())[0]:
                NewPremise = list(NewRule.values())[0]
                premise = list(rule.values())[0]
                if NewPremise == premise:
                    return self._Rules.index(rule)
        return False
    
    def add_rule(self, NewRule, NewRulesPath = None, ifSave = True):
        """Adding a new rule to the rule base
        
        Saving the new rules in excel on a path or not
        
        Args:
            NewRule(dict): The rule to be added
            NewRulesPath(str): The path for the new rule base file saving
            ifSave(bool): Whether or not to save the new rule base file
        
        Returns:
            (bool): Adding seuccessfully or not
        """
        if NewRulesPath == None:
            NewRulesPath = self._RulesPath
        if not self._if_exist_rule(NewRule):
            self._Rules.append(NewRule)
            print('Add the new rule successfully~')
            if ifSave:
                self._save_rules(NewRulesPath)
            return True
        else:
            print('The new rule has been existed~')
            return False
        
    def delete_rule(self, OldRule, NewRulesPath = None, ifSave = True):
        """Dropping a rule from the rule base
        
        Saving the new rules in excel on a path or not
        
        Args:
            OldRule(dict): The rule to be dropped
            NewRulesPath(str): The path for the new rule base file saving
            ifSave(bool): Whether or not to save the new rule base file
        
        Returns:
            None
        """
        if NewRulesPath == None:
            NewRulesPath = self._RulesPath
        if self._if_exist_rule(OldRule):
            self._Rules.pop(self._if_exist_rule(OldRule))
            print('The old rule has been deleted~')
            if ifSave:
                self._save_rules(NewRulesPath)
        else:
            print('The old rule does not exist~')
            
    def modify_rule(self, OldRule, NewRule, NewRulesPath = None, ifSave = True):
        """Modifying a rule in the rule base
        
        Saving the new rules in excel on a path or not
        
        Args:
            OldRule(dict): The rule before changing
            NewRule(dict): The rule after changing
            NewRulesPath(str): The path for the new rule base file saving
            ifSave(bool): Whether or not to save the new rule base file
        
        Returns:
            None
        """
        if NewRulesPath == None:
            NewRulesPath = self._RulesPath
        if self._if_exist_rule(OldRule):
            if not self._if_exist_rule(NewRule):
                self.delete_rule(OldRule, NewRulesPath, ifSave)
                self.add_rule(NewRule, NewRulesPath, ifSave)
            else:
                print('The new rule has been existed~')
        else:
            print('The old rule does not exist~')